from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.cache import never_cache
from django.contrib import messages
from django.db.models import Sum, Q, Count
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from .models import Reception, SortieStockProduction, Profile, Fournisseur, OuvertureProduction, LotReception
from .forms import ReceptionForm, ReceptionUpdateForm, SortieStockProductionForm, OuvertureProductionForm, ProfileCreateForm, ProfileUpdateForm
from .utils import lot_expiration_info
from . import mixins_fixed
from .decorators import magasin_required, admin_required, production_required

class OuvertureProductionCreateView(mixins_fixed.ProductionAccessMixin, CreateView):
    model = OuvertureProduction
    form_class = OuvertureProductionForm
    template_name = 'ouverture_production_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        return kwargs
    
    def form_valid(self, form):
        form.instance.profile = self.request.user.profile
        messages.success(self.request, 'Ouverture créée avec succès!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('production:index')

class OuvertureProductionListView(mixins_fixed.ProductionAccessMixin, ListView):
    model = OuvertureProduction
    template_name = 'ouverture_production_list.html'
    context_object_name = 'ouvertures'
    paginate_by = 50

    def get_queryset(self):
        return OuvertureProduction.objects.select_related('profile').order_by('-date_heure_ouverture')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        for ouverture in context['ouvertures']:
            ouverture.expiration_info = lot_expiration_info(ouverture.numero_lot)
        return context

@never_cache
def landing(request):
    return render(request, "landing.html")

@never_cache
def login_view(request):
    # If already authenticated, redirect directly to appropriate interface
    if request.user.is_authenticated:
        try:
            profile_type = request.user.profile.type_operateur
        except Profile.DoesNotExist:
            profile_type = None
        if profile_type == 'magasin':
            return redirect(reverse('dashboard'))
        if profile_type == 'production':
            return redirect(reverse('production:data_performance'))
        if profile_type == 'admin':
            next_page = request.GET.get('next', '')
            if next_page and 'magasin' in next_page:
                return redirect(reverse('dashboard'))
            return redirect(reverse('production:data_performance'))

    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)

            try:
                profile_type = user.profile.type_operateur
            except Profile.DoesNotExist:
                messages.error(request, "Profil manquant - contactez l'admin")
                return redirect('login')

            next_page = request.POST.get('next') or request.GET.get('next')

            # Admin: redirect based on which button was clicked
            if profile_type == 'admin':
                if next_page and 'magasin' in next_page:
                    return redirect(reverse('dashboard'))
                return redirect(reverse('production:data_performance'))

            # Magasin: always magasin interface
            if profile_type == 'magasin':
                return redirect(reverse('dashboard'))

            # Production: always production interface
            if profile_type == 'production':
                return redirect(reverse('production:data_performance'))

            return redirect(reverse('login'))
        else:
            messages.error(request, 'Identifiants invalides.')
    next_page = request.GET.get('next')
    return render(request, "login.html", {'next': next_page})

@never_cache
@production_required
def data_performance(request):
    from django.db.models.functions import TruncMonth
    from django.db.models import Count, Max
    from django.utils import timezone as tz
    from datetime import timedelta
    import json, calendar

    today = tz.localdate()
    now   = tz.localtime()

    # ── Période filter ──────────────────────────────────────────
    periode = request.GET.get('periode', 'all')
    qs = OuvertureProduction.objects.all()
    if periode == 'today':
        qs = qs.filter(date_heure_ouverture__date=today)
    elif periode == '7j':
        qs = qs.filter(date_heure_ouverture__date__gte=today - timedelta(days=7))
    elif periode == '30j':
        qs = qs.filter(date_heure_ouverture__date__gte=today - timedelta(days=30))

    # ── Opérateur filter ────────────────────────────────────────
    is_admin = request.user.profile.type_operateur == 'admin'
    operateur_filter = request.GET.get('operateur', '')
    operateurs_list  = Profile.objects.filter(type_operateur='production').order_by('nom')
    if operateur_filter:
        qs = qs.filter(profile_id=operateur_filter)

    # ── KPIs aujourd'hui ────────────────────────────────────────
    qs_today = OuvertureProduction.objects.filter(date_heure_ouverture__date=today)
    today_total = int(qs_today.aggregate(t=Sum('quantite'))['t'] or 0)
    today_count = qs_today.count()

    # ── KPIs période ────────────────────────────────────────────
    period_total = int(qs.aggregate(t=Sum('quantite'))['t'] or 0)
    period_count = qs.count()

    # ── Top shift & produit ─────────────────────────────────────
    top_shift_row   = qs.values('shift').annotate(t=Sum('quantite')).order_by('-t').first()
    top_product_row = qs.values('nom_produit').annotate(t=Sum('quantite')).order_by('-t').first()
    top_shift   = f"Shift {top_shift_row['shift']}" if top_shift_row else '—'
    top_product = top_product_row['nom_produit']   if top_product_row else '—'

    # ── Consommation mensuelle (6 mois) ─────────────────────────
    def prev_month(y, m):
        m -= 1
        if m == 0: m, y = 12, y - 1
        return y, m

    def month_bounds(y, m):
        last = calendar.monthrange(y, m)[1]
        from datetime import date
        return date(y, m, 1), date(y, m, last)

    MONTH_FR = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
    monthly_labels, monthly_data = [], []
    y, m = today.year, today.month
    months_seq = []
    for _ in range(6):
        months_seq.insert(0, (y, m))
        y, m = prev_month(y, m)
    for y, m in months_seq:
        s, e = month_bounds(y, m)
        qty = int(OuvertureProduction.objects.filter(
            date_heure_ouverture__date__gte=s,
            date_heure_ouverture__date__lte=e
        ).aggregate(t=Sum('quantite'))['t'] or 0)
        monthly_labels.append(MONTH_FR[m - 1])
        monthly_data.append(qty)

    # ── Par shift ───────────────────────────────────────────────
    shift_rows = list(qs.values('shift').annotate(total=Sum('quantite'), ops=Count('id')).order_by('shift'))

    # ── Par ligne ───────────────────────────────────────────────
    by_line = list(qs.values('ligne_production').annotate(total=Sum('quantite'), ops=Count('id')).order_by('-total')[:8])
    max_line = max((r['total'] for r in by_line), default=1) or 1

    # ── Par produit ─────────────────────────────────────────────
    by_product = list(qs.values('nom_produit').annotate(total=Sum('quantite'), ops=Count('id')).order_by('-total')[:8])
    max_product = max((r['total'] for r in by_product), default=1) or 1

    # ── Par opérateur ───────────────────────────────────────────
    by_op = list(qs.values('profile__nom','profile__prenom','profile__matricule').annotate(total=Sum('quantite'), ops=Count('id')).order_by('-total'))
    max_op = max((r['total'] for r in by_op), default=1) or 1
    for r in by_op:
        r['pct'] = round(r['total'] / max_op * 100)

    # ── Calculateur produit ─────────────────────────────────────
    produit_selected  = request.GET.get('produit', '')
    produits_existants = list(OuvertureProduction.objects.values_list('nom_produit', flat=True).distinct().order_by('nom_produit'))
    produit_total = 0
    if produit_selected:
        produit_total = int(qs.filter(nom_produit=produit_selected).aggregate(t=Sum('quantite'))['t'] or 0)

    context = {
        'periode_selected':        periode,
        'analytics_has_data':      qs.exists(),
        'realtime_now':            now,
        # KPIs
        'today_total':             today_total,
        'today_count':             today_count,
        'period_total':            period_total,
        'period_count':            period_count,
        'top_shift':               top_shift,
        'top_product':             top_product,
        # Charts JSON
        'monthly_labels':          json.dumps(monthly_labels),
        'monthly_data':            json.dumps(monthly_data),
        'shift_labels':            json.dumps([f"Shift {r['shift']}" for r in shift_rows]),
        'shift_data':              json.dumps([int(r['total']) for r in shift_rows]),
        'line_labels':             json.dumps([r['ligne_production'] or '—' for r in by_line]),
        'line_data':               json.dumps([int(r['total']) for r in by_line]),
        'product_labels':          json.dumps([r['nom_produit'] or '—' for r in by_product]),
        'product_data':            json.dumps([int(r['total']) for r in by_product]),
        # Tables
        'shift_rows':              shift_rows,
        'by_line':                 by_line,
        'max_line':                max_line,
        'by_product':              by_product,
        'max_product':             max_product,
        'by_op':                   by_op,
        # Calculateur
        'produits_existants':      produits_existants,
        'produit_selected':        produit_selected,
        'produit_total':           produit_total,
        # Admin
        'is_admin':                is_admin,
        'operateurs_list':         operateurs_list,
        'operateur_filter_selected': operateur_filter,
    }
    return render(request, 'data_performance.html', context)

@never_cache
@production_required
def profile_production(request):
    profiles = Profile.objects.filter(type_operateur__in=['production', 'admin']).order_by('nom')
    is_admin = hasattr(request.user, 'profile') and request.user.profile.type_operateur == 'admin'
    return render(request, 'profile_production.html', {'profiles': profiles, 'is_admin': is_admin})

@never_cache
@production_required
def profile_production_detail(request, pk):
    profile = get_object_or_404(Profile, pk=pk)
    ouvertures = OuvertureProduction.objects.filter(profile=profile).order_by('-date_heure_ouverture')[:20]
    return render(request, 'profile_production_detail.html', {'profile': profile, 'ouvertures': ouvertures})

@never_cache
@login_required
@csrf_protect
@production_required
def production(request):
    receptions = Reception.objects.select_related('fournisseur', 'profile').order_by('date_expiration')
    for r in receptions:
        r.status = 'ok'
        exp_info = lot_expiration_info(r.lot_reference[4:] if r.lot_reference else '')
        if exp_info['status'] == 'expired':
            r.status = 'expired'
        elif exp_info['status'] == 'warning':
            r.status = 'warning'
    total_stock = Reception.objects.aggregate(Sum('quantite'))['quantite__sum'] or 0
    
    profiles = Profile.objects.filter(type_operateur__in=['production', 'magasin']).order_by('nom')
    
    form = OuvertureProductionForm()
    if request.method == 'POST':
        form = OuvertureProductionForm(request.POST)
        if form.is_valid():
            try:
                instance = form.save(commit=False)
                instance.profile = request.user.profile
                instance.save()
                messages.success(request, 'Ouverture de production enregistrée avec succès!')
                return redirect('production:index')
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'enregistrement: {str(e)}')
        else:
            messages.error(request, 'Formulaire invalide. Vérifiez les champs.')
    
    context = {
        'receptions': receptions,
        'total_stock': total_stock,
        'profiles': profiles,
        'form': form,
    }
    return render(request, 'production.html', context)

@magasin_required
def reception_list(request):
    query = request.GET.get('q', '')
    receptions = Reception.objects.all()
    if query:
        receptions = receptions.filter(
            Q(profile__matricule__icontains=query) |
            Q(fournisseur__nom__icontains=query) |
            Q(lot_code1__icontains=query) |
            Q(lot_code2__icontains=query) |
            Q(lot_code3__icontains=query) |
            Q(lot_code4__icontains=query) |
            Q(lot_code5__icontains=query)
        )
    receptions = receptions.order_by('-date_reception')
    context = {'receptions': receptions}
    return render(request, "reception_list.html", context)

@never_cache
@production_required
def stock_actuel_production(request):
    from django.utils import timezone as tz
    today = tz.localdate()
    receptions = Reception.objects.select_related('fournisseur', 'profile').order_by('date_expiration')
    for r in receptions:
        days_left = (r.date_expiration - today).days
        if days_left < 0:
            r.status = 'expired'
        elif days_left <= 7:
            r.status = 'warning'
        else:
            r.status = 'ok'
    context = {'receptions': receptions, 'total_stock': Reception.objects.aggregate(Sum('quantite'))['quantite__sum'] or 0}
    return render(request, "stock_actuel.html", context)

@magasin_required
def stock_actuel(request):
    from django.utils import timezone as tz
    from datetime import timedelta
    today = tz.localdate()
    receptions = Reception.objects.select_related('fournisseur', 'profile').order_by('date_expiration')
    for r in receptions:
        days_left = (r.date_expiration - today).days
        if days_left < 0:
            r.status = 'expired'
        elif days_left <= 7:
            r.status = 'warning'
        else:
            r.status = 'ok'
    quantite_expiree = round(float(Reception.objects.filter(date_expiration__lt=today).aggregate(t=Sum('quantite'))['t'] or 0), 3)
    quantite_warning = round(float(Reception.objects.filter(
        date_expiration__gte=today,
        date_expiration__lte=today + timedelta(days=7)
    ).aggregate(t=Sum('quantite'))['t'] or 0), 3)
    context = {
        'receptions': receptions,
        'total_stock': Reception.objects.aggregate(Sum('quantite'))['quantite__sum'] or 0,
        'quantite_expiree': quantite_expiree,
        'quantite_warning': quantite_warning,
    }
    return render(request, "stock_actuel.html", context)

@never_cache
@production_required
def historique(request):
    ouvertures = OuvertureProduction.objects.select_related('profile').order_by('-date_heure_ouverture')
    for o in ouvertures:
        o.expiration_info = lot_expiration_info(o.numero_lot)
    context = {'ouvertures': ouvertures}
    return render(request, "historique.html", context)

@production_required
def ouverture_detail(request, pk):
    ouverture = get_object_or_404(OuvertureProduction, pk=pk)
    ouverture.expiration_info = lot_expiration_info(ouverture.numero_lot)
    return render(request, "ouverture_detail.html", {'ouverture': ouverture})

@admin_required
def ouverture_update(request, pk):
    ouverture = get_object_or_404(OuvertureProduction, pk=pk)
    if request.method == 'POST':
        form = OuvertureProductionForm(request.POST, instance=ouverture)
        if form.is_valid():
            form.save()
            messages.success(request, 'Enregistrement modifié avec succès.')
            return redirect('production:historique')
    else:
        form = OuvertureProductionForm(instance=ouverture)
    # Bootstrap classes
    for fname, fld in form.fields.items():
        if hasattr(fld.widget, 'attrs'):
            if 'class' not in fld.widget.attrs:
                fld.widget.attrs['class'] = 'form-control'
    return render(request, "ouverture_update_form.html", {'form': form, 'ouverture': ouverture})

@admin_required
def ouverture_delete(request, pk):
    ouverture = get_object_or_404(OuvertureProduction, pk=pk)
    if request.method == 'POST':
        ouverture.delete()
        messages.success(request, 'Enregistrement supprimé.')
        return redirect('production:historique')
    return render(request, "ouverture_confirm_delete.html", {'ouverture': ouverture})

@magasin_required
def dashboard(request):
    import json, calendar
    from datetime import date, timedelta
    from django.utils import timezone

    today        = timezone.localdate()
    first_of_month = today.replace(day=1)
    profile      = getattr(request.user, 'profile', None)

    # ── Helper: month bounds ────────────────────────────────────
    def month_bounds(y, m):
        last = calendar.monthrange(y, m)[1]
        return date(y, m, 1), date(y, m, last)

    def prev_month(y, m):
        m -= 1
        if m == 0:
            m, y = 12, y - 1
        return y, m

    # ── Stats globales ──────────────────────────────────────────
    total_recu   = float(Reception.objects.aggregate(t=Sum('quantite'))['t'] or 0)
    total_envoye = float(SortieStockProduction.objects.aggregate(t=Sum('quantite'))['t'] or 0)
    # Stock disponible = reçu − transféré uniquement (les lots expirés ne sont PAS déduits)
    stock_dispo  = total_recu - total_envoye
    utilisation_pct = round(total_envoye / total_recu * 100, 1) if total_recu else 0

    total_receptions = Reception.objects.count()
    total_transfers  = SortieStockProduction.objects.count()

    # Ce mois
    receptions_mois = Reception.objects.filter(date_reception__date__gte=first_of_month).count()
    transfers_mois  = SortieStockProduction.objects.filter(date_heure__date__gte=first_of_month).count()

    # Mois précédent (trend)
    py, pm = prev_month(today.year, today.month)
    ps, pe = month_bounds(py, pm)
    prev_rec   = Reception.objects.filter(date_reception__date__gte=ps, date_reception__date__lte=pe).count()
    prev_trans = SortieStockProduction.objects.filter(date_heure__date__gte=ps, date_heure__date__lte=pe).count()

    def trend(cur, prev):
        if prev == 0:
            return None, None
        pct = round((cur - prev) / prev * 100, 1)
        return pct, 'up' if pct >= 0 else 'down'

    trend_rec_pct,   trend_rec_dir   = trend(receptions_mois, prev_rec)
    trend_trans_pct, trend_trans_dir = trend(transfers_mois,  prev_trans)

    # ── Alertes expiration ───────────────────────────────────────
    lots_expires_count = Reception.objects.filter(date_expiration__lt=today).count()
    lots_warning_count = Reception.objects.filter(
        date_expiration__gte=today,
        date_expiration__lte=today + timedelta(days=7)
    ).count()
    quantite_expiree = round(float(Reception.objects.filter(date_expiration__lt=today).aggregate(t=Sum('quantite'))['t'] or 0), 3)
    quantite_warning = round(float(Reception.objects.filter(
        date_expiration__gte=today,
        date_expiration__lte=today + timedelta(days=7)
    ).aggregate(t=Sum('quantite'))['t'] or 0), 3)
    lots_expires_list = Reception.objects.filter(
        date_expiration__lt=today
    ).select_related('fournisseur').order_by('date_expiration')[:5]
    lots_warning_list = Reception.objects.filter(
        date_expiration__gte=today,
        date_expiration__lte=today + timedelta(days=7)
    ).select_related('fournisseur').order_by('date_expiration')[:5]

    # ── Graphique barres : 6 derniers mois ──────────────────────
    MONTH_FR = ['Jan','Fév','Mar','Avr','Mai','Jun','Jul','Aoû','Sep','Oct','Nov','Déc']
    labels, data_rec, data_trans = [], [], []
    y, m = today.year, today.month
    months = []
    for _ in range(6):
        months.insert(0, (y, m))
        y, m = prev_month(y, m)

    for y, m in months:
        s, e = month_bounds(y, m)
        rq = float(Reception.objects.filter(date_reception__date__gte=s, date_reception__date__lte=e).aggregate(t=Sum('quantite'))['t'] or 0)
        tq = float(SortieStockProduction.objects.filter(date_heure__date__gte=s, date_heure__date__lte=e).aggregate(t=Sum('quantite'))['t'] or 0)
        labels.append(MONTH_FR[m - 1])
        data_rec.append(round(rq, 3))
        data_trans.append(round(tq, 3))

    # ── Activité récente ─────────────────────────────────────────
    recent_receptions = Reception.objects.select_related('profile', 'fournisseur').order_by('-date_reception')[:8]
    recent_transfers  = SortieStockProduction.objects.select_related('profile').order_by('-date_heure')[:8]

    context = {
        'stock_dispo':        round(stock_dispo, 3),
        'total_recu':         round(total_recu, 3),
        'total_envoye':       round(total_envoye, 3),
        'utilisation_pct':    utilisation_pct,
        'total_receptions':   total_receptions,
        'total_transfers':    total_transfers,
        'receptions_mois':    receptions_mois,
        'transfers_mois':     transfers_mois,
        'trend_rec_pct':      trend_rec_pct,
        'trend_rec_dir':      trend_rec_dir,
        'trend_trans_pct':    trend_trans_pct,
        'trend_trans_dir':    trend_trans_dir,
        'lots_expires_count': lots_expires_count,
        'lots_warning_count': lots_warning_count,
        'quantite_expiree':   quantite_expiree,
        'quantite_warning':   quantite_warning,
        'lots_expires_list':  lots_expires_list,
        'lots_warning_list':  lots_warning_list,
        'recent_receptions':  recent_receptions,
        'recent_transfers':   recent_transfers,
        'chart_labels':       json.dumps(labels),
        'chart_rec':          json.dumps(data_rec),
        'chart_trans':        json.dumps(data_trans),
        'donut_dispo':        round(stock_dispo, 3),
        'donut_envoye':       round(total_envoye, 3),
        'current_profile':    profile,
    }
    return render(request, 'dashboard.html', context)

@magasin_required
def transfert_create(request):
    import json
    from django.db import transaction
    from .utils import lot_exists, lot_not_expired, lot_total_received, lot_total_sent_to_production, canonical_lot_identifier
    from django.utils import timezone as tz

    # Build lot choices with real disponible
    from .utils import all_lot_options_for_sortie, lot_total_received, lot_total_sent_to_production
    from django.utils import timezone as tz
    today = tz.localdate()
    lot_choices = []
    for option in all_lot_options_for_sortie(today):
        lot_id = option['lot_id']
        dispo = float(lot_total_received(lot_id) - lot_total_sent_to_production(lot_id))
        if dispo <= 0:
            continue
        date_exp = option['date_expiration'].strftime('%d/%m/%Y')
        jours = option.get('jours_restant')
        statut = ''
        if jours is not None and jours < 0:
            statut = ' | ⛔ EXPIRÉ'
        elif jours is not None and jours <= 7:
            statut = f' | ⚠ EXP ≤7j ({jours}j)'
        label = f"{lot_id} | Dispo: {dispo:.3f} kg | Exp: {date_exp}{statut}"
        lot_choices.append({'value': lot_id, 'label': label, 'disponible': dispo})

    form_ref = SortieStockProductionForm(current_user=request.user)
    profile_choices = [{'value': str(p.pk), 'label': str(p)} for p in form_ref.fields['profile'].queryset]

    if request.method == 'POST':
        profile_id = request.POST.get('profile')
        errors = []
        lots_data = []

        # Parse dynamic lots from POST: numero_lot_0, quantite_0, ...
        i = 0
        while f'numero_lot_{i}' in request.POST:
            lot_id = request.POST.get(f'numero_lot_{i}', '').strip()
            try:
                quantite = float(request.POST.get(f'quantite_{i}', 0))
            except ValueError:
                quantite = 0
            if lot_id or quantite:
                lots_data.append({'lot_id': lot_id, 'quantite': quantite, 'index': i + 1})
            i += 1

        if not profile_id:
            errors.append("Veuillez sélectionner un opérateur.")
        if not lots_data:
            errors.append("Ajoutez au moins un lot.")

        # Resolve canonical IDs and basic checks
        today = tz.localdate()
        for entry in lots_data:
            lot_id = canonical_lot_identifier(entry['lot_id'])
            if not lot_id:
                errors.append(f"Lot {entry['index']}: ID invalide.")
                continue
            entry['canonical'] = lot_id
            if not lot_exists(lot_id):
                errors.append(f"Lot {entry['index']}: n'existe pas en stock.")
            elif not lot_not_expired(lot_id, today):
                errors.append(f"Lot {entry['index']}: expiré.")
            elif entry['quantite'] <= 0:
                errors.append(f"Lot {entry['index']}: quantité invalide.")

        # Aggregate quantities per lot and check stock (handles same lot in multiple rows)
        if not errors:
            from collections import defaultdict
            lot_totals = defaultdict(float)
            for entry in lots_data:
                lot_totals[entry.get('canonical', '')] += entry['quantite']
            for lot_id, total_needed in lot_totals.items():
                if not lot_id:
                    continue
                dispo = float(lot_total_received(lot_id) - lot_total_sent_to_production(lot_id))
                if total_needed > dispo:
                    errors.append(f"Lot {lot_id}: quantité totale insuffisante (demandé: {total_needed:.3f} kg, disponible: {dispo:.3f} kg).")

        if not errors:
            try:
                profile = Profile.objects.get(pk=profile_id)
                with transaction.atomic():
                    for entry in lots_data:
                        SortieStockProduction.objects.create(
                            profile=profile,
                            numero_lot=entry['canonical'],
                            quantite=entry['quantite'],
                        )
                messages.success(request, f"{len(lots_data)} transfert(s) enregistré(s) avec succès.")
                return redirect('transfert_list')
            except Profile.DoesNotExist:
                errors.append("Opérateur introuvable.")

        context = {
            'lot_choices_json': json.dumps(lot_choices),
            'profile_choices': profile_choices,
            'errors': errors,
            'profile_selected': profile_id,
        }
        return render(request, 'sortie_stock_form.html', context)

    context = {
        'lot_choices_json': json.dumps(lot_choices),
        'profile_choices': profile_choices,
        'errors': [],
        'profile_selected': '',
    }
    return render(request, 'sortie_stock_form.html', context)

@magasin_required
def transfert_list(request):
    query = request.GET.get('q', '')
    sorties = SortieStockProduction.objects.select_related('profile').all()
    if query:
        sorties = sorties.filter(
            Q(profile__matricule__icontains=query) |
            Q(numero_lot__icontains=query)
        )
    sorties = sorties.order_by('-date_heure')[:50]
    for s in sorties:
        s.expiration_info = lot_expiration_info(s.numero_lot)
    context = {'sorties': sorties}
    return render(request, "transfert_list.html", context)

class ReceptionCreateView(mixins_fixed.MagasinAccessMixin, CreateView):
    model = Reception
    form_class = ReceptionForm
    template_name = 'reception_form.html'
    success_url = '/gestion_stock/reception/'

    def _parse_lots_from_post(self, post):
        from decimal import Decimal, InvalidOperation
        lot_count = int(post.get('lot_count', 0) or 0)
        lots = []
        errors = []
        for i in range(1, lot_count + 1):
            code = post.get(f'lot_code{i}', '').strip()
            qty_str = post.get(f'lot_quantite{i}', '').strip()
            if not code:
                continue
            try:
                qty = Decimal(qty_str)
            except (InvalidOperation, ValueError):
                errors.append(f"Quantité invalide pour le lot {i}.")
                continue
            if qty <= 0:
                errors.append(f"La quantité du lot {i} doit être > 0.")
                continue
            lots.append({'code': code, 'quantite': qty, 'ordre': i})
        return lots, errors

    def form_valid(self, form):
        from decimal import Decimal
        lots, errors = self._parse_lots_from_post(self.request.POST)
        if errors or not lots:
            for e in errors:
                form.add_error(None, e)
            if not lots:
                form.add_error(None, "Veuillez saisir au moins un lot.")
            return self.form_invalid(form)

        form.instance.profile = self.request.user.profile
        form.instance.quantite = sum(l['quantite'] for l in lots)
        response = super().form_valid(form)

        for lot_data in lots:
            LotReception.objects.create(
                reception=self.object,
                code=lot_data['code'],
                quantite=lot_data['quantite'],
                ordre=lot_data['ordre'],
            )
        return response

    def get_context_data(self, **kwargs):
        import json
        context = super().get_context_data(**kwargs)
        form = context['form']
        lots_data = []
        if form.is_bound:
            lot_count = int(form.data.get('lot_count', 0) or 0)
            for i in range(1, lot_count + 1):
                code = form.data.get(f'lot_code{i}', '').strip()
                qty = form.data.get(f'lot_quantite{i}', '')
                if code:
                    lots_data.append({'code': code, 'qty': qty})
        context['lots_restore_json'] = json.dumps(lots_data)
        return context

@magasin_required
def profile_list(request):
    profiles = Profile.objects.filter(type_operateur__in=['magasin', 'admin']).order_by('nom')
    is_admin = hasattr(request.user, 'profile') and request.user.profile.type_operateur == 'admin'
    return render(request, 'profile.html', {'profiles': profiles, 'is_admin': is_admin})

@magasin_required
def profile_detail(request, pk):
    profile = get_object_or_404(Profile, pk=pk)
    receptions = Reception.objects.filter(profile=profile).select_related('fournisseur').order_by('-date_reception')[:20]
    transferts = SortieStockProduction.objects.filter(profile=profile).order_by('-date_heure')[:20]
    return render(request, 'profile_detail.html', {
        'profile': profile,
        'receptions': receptions,
        'transferts': transferts,
    })

@magasin_required
def profile_add(request):
    if not (hasattr(request.user, 'profile') and request.user.profile.type_operateur == 'admin'):
        messages.error(request, "Accès réservé aux administrateurs.")
        return redirect('profile')
    if request.method == 'POST':
        form = ProfileCreateForm(request.POST, allowed_types=['magasin', 'admin'])
        if form.is_valid():
            matricule = form.cleaned_data['matricule']
            user = User.objects.create_user(username=matricule, password=form.cleaned_data['password'])
            Profile.objects.create(
                user=user,
                matricule=matricule,
                nom=form.cleaned_data['nom'],
                prenom=form.cleaned_data['prenom'],
                type_operateur=form.cleaned_data['type_operateur'],
            )
            messages.success(request, f"Profile « {matricule} » créé avec succès.")
            return redirect('profile')
    else:
        form = ProfileCreateForm(allowed_types=['magasin', 'admin'])
    return render(request, 'profile_add.html', {
        'form': form,
        'back_url': reverse('profile'),
        'title': 'Ajouter un Magasinier',
        'context_type': 'magasin',
    })

@production_required
def profile_production_add(request):
    if not (hasattr(request.user, 'profile') and request.user.profile.type_operateur == 'admin'):
        messages.error(request, "Accès réservé aux administrateurs.")
        return redirect('production:profile')
    if request.method == 'POST':
        form = ProfileCreateForm(request.POST, allowed_types=['production'])
        if form.is_valid():
            matricule = form.cleaned_data['matricule']
            user = User.objects.create_user(username=matricule, password=form.cleaned_data['password'])
            Profile.objects.create(
                user=user,
                matricule=matricule,
                nom=form.cleaned_data['nom'],
                prenom=form.cleaned_data['prenom'],
                type_operateur=form.cleaned_data['type_operateur'],
            )
            messages.success(request, f"Profile « {matricule} » créé avec succès.")
            return redirect('production:profile')
    else:
        form = ProfileCreateForm(allowed_types=['production'])
    return render(request, 'profile_add.html', {
        'form': form,
        'back_url': reverse('production:profile'),
        'title': 'Ajouter un Opérateur Production',
        'context_type': 'production',
    })

@magasin_required
def profile_update(request, pk):
    if not (hasattr(request.user, 'profile') and request.user.profile.type_operateur == 'admin'):
        messages.error(request, "Accès réservé aux administrateurs.")
        return redirect('profile')
    profile = get_object_or_404(Profile, pk=pk)
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=profile, allowed_types=['magasin', 'admin'])
        if form.is_valid():
            old_matricule = profile.matricule
            profile.matricule = form.cleaned_data['matricule']
            profile.nom = form.cleaned_data['nom']
            profile.prenom = form.cleaned_data['prenom']
            profile.type_operateur = form.cleaned_data['type_operateur']
            profile.save()
            if form.cleaned_data.get('password'):
                profile.user.set_password(form.cleaned_data['password'])
                profile.user.save()
            messages.success(request, f"Profile « {profile.matricule} » modifié avec succès.")
            return redirect('profile')
    else:
        form = ProfileUpdateForm(instance=profile, allowed_types=['magasin', 'admin'])
    return render(request, 'profile_update.html', {
        'form': form,
        'profile': profile,
        'back_url': reverse('profile'),
        'title': f'Modifier — {profile.matricule}',
        'context_type': 'magasin',
    })

@production_required
def profile_production_update(request, pk):
    if not (hasattr(request.user, 'profile') and request.user.profile.type_operateur == 'admin'):
        messages.error(request, "Accès réservé aux administrateurs.")
        return redirect('production:profile')
    profile = get_object_or_404(Profile, pk=pk)
    if request.method == 'POST':
        form = ProfileUpdateForm(request.POST, instance=profile, allowed_types=['production'])
        if form.is_valid():
            profile.matricule = form.cleaned_data['matricule']
            profile.nom = form.cleaned_data['nom']
            profile.prenom = form.cleaned_data['prenom']
            profile.type_operateur = form.cleaned_data['type_operateur']
            profile.save()
            if form.cleaned_data.get('password'):
                profile.user.set_password(form.cleaned_data['password'])
                profile.user.save()
            messages.success(request, f"Profile « {profile.matricule} » modifié avec succès.")
            return redirect('production:profile')
    else:
        form = ProfileUpdateForm(instance=profile, allowed_types=['production'])
    return render(request, 'profile_update.html', {
        'form': form,
        'profile': profile,
        'back_url': reverse('production:profile'),
        'title': f'Modifier — {profile.matricule}',
        'context_type': 'production',
    })

@magasin_required
def reception_detail(request, pk):
    reception = get_object_or_404(Reception, pk=pk)
    lots = [
        {
            'index': lot.ordre,
            'canonical': f"LOT-{reception.id}-{lot.ordre:02d}",
            'code': lot.code,
            'quantite': lot.quantite,
        }
        for lot in reception.lots.all()
    ]
    context = {
        'reception': reception,
        'lots': lots,
    }
    return render(request, 'reception_detail.html', context)

@magasin_required
def reception_export_excel(request, pk):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse

    reception = get_object_or_404(Reception, pk=pk)
    lots = [
        {
            'index': lot.ordre,
            'canonical': f"LOT-{reception.id}-{lot.ordre:02d}",
            'code': lot.code,
            'quantite': lot.quantite,
        }
        for lot in reception.lots.all()
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Réception {reception.id}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="17A2B8")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Titre
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Fiche de Réception N° {reception.id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center
    ws["A1"].fill = PatternFill("solid", fgColor="E0F7FA")

    # Infos générales
    ws["A3"] = "Opérateur"
    ws["B3"] = f"{reception.profile.nom} {reception.profile.prenom} ({reception.profile.matricule})"
    ws["A4"] = "Fournisseur"
    ws["B4"] = reception.fournisseur.nom
    ws["A5"] = "Date réception"
    ws["B5"] = reception.date_reception.strftime("%d/%m/%Y %H:%M")
    ws["A6"] = "Date expiration"
    ws["B6"] = reception.date_expiration.strftime("%d/%m/%Y") if reception.date_expiration else ""
    ws["A7"] = "Quantité totale (kg)"
    ws["B7"] = float(reception.quantite) if reception.quantite else 0

    for row in range(3, 8):
        ws[f"A{row}"].font = Font(bold=True)

    # En-têtes tableau
    headers = ["#", "ID Canonique", "Code Lot réel", "Quantité (kg)"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    # Données lots
    for row_idx, lot in enumerate(lots, start=10):
        ws.cell(row=row_idx, column=1, value=lot['index']).border = border
        ws.cell(row=row_idx, column=2, value=lot['canonical']).border = border
        ws.cell(row=row_idx, column=3, value=lot['code']).border = border
        qty_cell = ws.cell(row=row_idx, column=4, value=float(lot['quantite']) if lot['quantite'] else None)
        qty_cell.border = border
        qty_cell.alignment = Alignment(horizontal="right")

    # Ligne total
    total_row = 10 + len(lots)
    ws.cell(row=total_row, column=3, value="Total").font = Font(bold=True)
    ws.cell(row=total_row, column=3).alignment = Alignment(horizontal="right")
    ws.cell(row=total_row, column=4, value=float(reception.quantite) if reception.quantite else 0).font = Font(bold=True)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="reception_{reception.id}.xlsx"'
    wb.save(response)
    return response

class ReceptionUpdateView(mixins_fixed.AdminRequiredMixin, UpdateView):
    model = Reception
    form_class = ReceptionUpdateForm
    template_name = 'reception_update_form.html'
    success_url = '/gestion_stock/reception/'

    def _parse_lots_from_post(self, post):
        from decimal import Decimal, InvalidOperation
        lot_count = int(post.get('lot_count', 0) or 0)
        lots = []
        errors = []
        for i in range(1, lot_count + 1):
            code = post.get(f'lot_code{i}', '').strip()
            qty_str = post.get(f'lot_quantite{i}', '').strip()
            if not code:
                continue
            try:
                qty = Decimal(qty_str)
            except (InvalidOperation, ValueError):
                errors.append(f"Quantité invalide pour le lot {i}.")
                continue
            if qty <= 0:
                errors.append(f"La quantité du lot {i} doit être > 0.")
                continue
            lots.append({'code': code, 'quantite': qty, 'ordre': i})
        return lots, errors

    def form_valid(self, form):
        lots, errors = self._parse_lots_from_post(self.request.POST)
        if errors or not lots:
            for e in errors:
                form.add_error(None, e)
            if not lots:
                form.add_error(None, "Veuillez saisir au moins un lot.")
            return self.form_invalid(form)

        form.instance.quantite = sum(l['quantite'] for l in lots)
        response = super().form_valid(form)

        self.object.lots.all().delete()
        for lot_data in lots:
            LotReception.objects.create(
                reception=self.object,
                code=lot_data['code'],
                quantite=lot_data['quantite'],
                ordre=lot_data['ordre'],
            )
        return response

    def get_context_data(self, **kwargs):
        import json
        context = super().get_context_data(**kwargs)
        reception = self.object
        if self.request.method == 'POST':
            lot_count = int(self.request.POST.get('lot_count', 0) or 0)
            lots_data = []
            for i in range(1, lot_count + 1):
                code = self.request.POST.get(f'lot_code{i}', '').strip()
                qty = self.request.POST.get(f'lot_quantite{i}', '')
                if code:
                    lots_data.append({'code': code, 'qty': qty})
        else:
            lots_data = [
                {'code': lot.code, 'qty': str(lot.quantite)}
                for lot in reception.lots.all()
            ]
        context['lots_restore_json'] = json.dumps(lots_data)
        return context

@admin_required
def reception_delete(request, pk):
    reception = get_object_or_404(Reception, pk=pk)
    if request.method == 'POST':
        reception.delete()
        return redirect('reception_list')
    context = {'reception': reception}
    return render(request, 'reception_confirm_delete.html', context)

@magasin_required
def transfert_detail(request, pk):
    from .utils import lot_id_to_code, _parse_lot_identifier
    sortie = get_object_or_404(SortieStockProduction, pk=pk)
    exp_info = lot_expiration_info(sortie.numero_lot)
    lot_code_reel = lot_id_to_code(sortie.numero_lot) or sortie.numero_lot
    parsed = _parse_lot_identifier(sortie.numero_lot)
    reception = None
    if parsed:
        from .models import Reception as Rec
        reception = Rec.objects.filter(pk=parsed[0]).first()
    context = {
        'sortie': sortie,
        'expiration_info': exp_info,
        'lot_code_reel': lot_code_reel,
        'reception': reception,
    }
    return render(request, 'transfert_detail.html', context)

@admin_required
def transfert_update(request, pk):
    sortie = get_object_or_404(SortieStockProduction, pk=pk)
    if request.method == 'POST':
        form = SortieStockProductionForm(request.POST, instance=sortie, current_user=request.user)
        if form.is_valid():
            form.save()
            return redirect('transfert_list')
    else:
        form = SortieStockProductionForm(instance=sortie, current_user=request.user)
    # Add Bootstrap classes to widgets
    form.fields['profile'].widget.attrs.update({'class': 'form-select'})
    form.fields['numero_lot'].widget.attrs.update({'class': 'form-select'})
    form.fields['quantite'].widget.attrs.update({'class': 'form-control', 'step': '0.001', 'min': '0.001'})
    context = {'form': form, 'sortie': sortie}
    return render(request, 'transfert_update_form.html', context)

@admin_required
def transfert_delete(request, pk):
    sortie = get_object_or_404(SortieStockProduction, pk=pk)
    if request.method == 'POST':
        sortie.delete()
        return redirect('transfert_list')
    context = {'sortie': sortie}
    return render(request, 'transfert_confirm_delete.html', context)

@never_cache
@production_required
def nouvelle_production(request):
    form = OuvertureProductionForm()
    if request.method == 'POST':
        form = OuvertureProductionForm(request.POST)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.profile = request.user.profile
            instance.save()
            messages.success(request, 'Nouvelle production créée avec succès!')
            return redirect('production:index')
        else:
            messages.error(request, 'Formulaire invalide. Vérifiez les champs.')
    
    context = {'form': form}
    return render(request, 'ouverture_production_form.html', context)

